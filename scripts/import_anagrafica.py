#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Importa anagrafica titoli da file Excel nel database cedi.

Popola: collane, autori, titoli, titolo_autore, titolo_codifica.

Usage:
    python scripts/import_anagrafica.py [--instance INSTANCE] [--clear] [--file FILE]
"""

import sys
import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
from gnr.app.gnrapp import GnrApp

PROJECT_ROOT = Path(__file__).parent.parent


def leggi_excel(filepath):
    """Legge il foglio Master_Pulito e restituisce una lista di dict."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Master_Pulito']
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        record = dict(zip(headers, row))
        if record.get('Titolo'):
            rows.append(record)
    return rows


def raggruppa_per_titolo(rows):
    """Raggruppa righe per Titolo. Raccoglie autori distinti."""
    grouped = defaultdict(list)
    for row in rows:
        titolo = (row.get('Titolo') or '').strip()
        if titolo:
            grouped[titolo].append(row)
    return grouped


def importa(db, rows, clear=False):
    tbl_collana = db.table('cedi_base.collana')
    tbl_autore = db.table('cedi_base.autore')
    tbl_titolo = db.table('cedi_base.titolo')
    tbl_titolo_autore = db.table('cedi_base.titolo_autore')
    tbl_titolo_codifica = db.table('cedi_base.titolo_codifica')

    if clear:
        print("Pulizia tabelle esistenti...")
        db.rollback()
        for tbl_name in ('cedi_base_titolo_codifica', 'cedi_base_titolo_autore',
                         'cedi_base_titolo', 'cedi_base_autore', 'cedi_base_collana'):
            db.execute('TRUNCATE TABLE cedi_base.%s CASCADE' % tbl_name)
        db.commit()
        print("  Tabelle svuotate.")

    # Cache collane esistenti
    collane_cache = {}
    for r in tbl_collana.query(columns='$codice, $descrizione').fetch():
        collane_cache[r['descrizione'].upper()] = r['codice']

    # Cache autori esistenti
    autori_cache = {}
    for r in tbl_autore.query(columns='$id, $cognome, $nome').fetch():
        denominazione = ('%s %s' % (r['nome'] or '', r['cognome'] or '')).strip().upper()
        autori_cache[denominazione] = r['id']

    grouped = raggruppa_per_titolo(rows)
    tot = len(grouped)
    contatori = dict(collane=0, autori=0, titoli=0, titolo_autore=0, codifiche=0)

    print("Importazione %d titoli da %d righe Excel..." % (tot, len(rows)))

    for i, (titolo_nome, righe) in enumerate(grouped.items(), 1):
        prima_riga = righe[0]

        # Collana
        collana_nome = (prima_riga.get('Collana') or '').strip()
        collana_codice = None
        if collana_nome:
            collana_codice = collane_cache.get(collana_nome.upper())
            if not collana_codice:
                codice = collana_nome[:5].upper()
                rec = tbl_collana.newrecord(codice=codice, descrizione=collana_nome)
                tbl_collana.insert(rec)
                collane_cache[collana_nome.upper()] = codice
                collana_codice = codice
                contatori['collane'] += 1

        # Titolo
        isbn_titolo = None
        for r in righe:
            if r.get('ISBN'):
                isbn_titolo = str(r['ISBN'])
                break

        data_pub = prima_riga.get('Data_Uscita')
        rec_titolo = tbl_titolo.newrecord(
            titolo=titolo_nome,
            isbn=isbn_titolo,
            collana_codice=collana_codice,
            data_pubblicazione=data_pub,
            attivo=True
        )
        tbl_titolo.insert(rec_titolo)
        titolo_id = rec_titolo['id']
        contatori['titoli'] += 1

        # Autori (possono essere piu di uno per titolo)
        autori_visti = set()
        for riga in righe:
            autore_nome = (riga.get('Autore') or '').strip()
            if not autore_nome or autore_nome.upper() in autori_visti:
                continue
            autori_visti.add(autore_nome.upper())
            autore_id = autori_cache.get(autore_nome.upper())
            if not autore_id:
                parti = autore_nome.rsplit(' ', 1)
                if len(parti) == 2:
                    nome, cognome = parti[0], parti[1]
                else:
                    nome = ''
                    cognome = parti[0]
                rec = tbl_autore.newrecord(nome=nome, cognome=cognome)
                tbl_autore.insert(rec)
                autore_id = rec['id']
                autori_cache[autore_nome.upper()] = autore_id
                contatori['autori'] += 1
            rec_ta = tbl_titolo_autore.newrecord(
                titolo_id=titolo_id,
                autore_id=autore_id,
                quota_percentuale=100
            )
            tbl_titolo_autore.insert(rec_ta)
            contatori['titolo_autore'] += 1

        # Codifiche
        for riga in righe:
            isbn = riga.get('ISBN')
            asin = riga.get('ASIN')
            isbn_storytel = riga.get('ISBN_STORYTEL')
            if isbn:
                codice = str(isbn)
                tipo_codice = 'ISBN'
            elif asin:
                codice = str(asin)
                tipo_codice = 'ASIN'
            elif isbn_storytel:
                codice = str(isbn_storytel)
                tipo_codice = 'ISBN_STORYTEL'
            else:
                continue
            rec_cod = tbl_titolo_codifica.newrecord(
                titolo_id=titolo_id,
                codice=codice,
                tipo_codice=tipo_codice,
                tipo_prodotto=riga.get('Tipo_Prodotto'),
                tipo_stampa=riga.get('Tipo_Stampa'),
                prezzo=riga.get('Prezzo'),
                royalty_percentuale=riga.get('Royalty'),
                data_uscita=riga.get('Data_Uscita')
            )
            tbl_titolo_codifica.insert(rec_cod)
            contatori['codifiche'] += 1

        if i % 50 == 0:
            db.commit()
            print("  %d/%d titoli..." % (i, tot))

    db.commit()
    return contatori


def stampa_riepilogo(contatori):
    print("\n" + "=" * 50)
    print("RIEPILOGO IMPORTAZIONE")
    print("=" * 50)
    print("  Collane create:        %d" % contatori['collane'])
    print("  Autori creati:         %d" % contatori['autori'])
    print("  Titoli creati:         %d" % contatori['titoli'])
    print("  Legami titolo-autore:  %d" % contatori['titolo_autore'])
    print("  Codifiche create:      %d" % contatori['codifiche'])
    print("=" * 50)


def main():
    parser = argparse.ArgumentParser(description='Importa anagrafica titoli da Excel')
    parser.add_argument('--instance', default='cedipg', help='Nome istanza (default: cedipg)')
    parser.add_argument('--clear', action='store_true', help='Svuota tabelle prima di importare')
    parser.add_argument('--file', default=str(PROJECT_ROOT / 'DELRAI' / 'anagrafiche' / 'Anagrafica_titoli.xlsx'),
                        help='Path file Excel')
    args = parser.parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        print("Errore: file non trovato: %s" % filepath)
        sys.exit(1)

    print("Connessione a istanza: %s" % args.instance)
    instance_dir = PROJECT_ROOT / 'instances' / args.instance
    if not instance_dir.exists():
        print("Errore: directory istanza non trovata: %s" % instance_dir)
        sys.exit(1)

    app = GnrApp(str(instance_dir))
    db = app.db

    print("Verifica schema database...")
    db.checkDb(applyChanges=True)

    print("Lettura file: %s" % filepath)
    rows = leggi_excel(filepath)
    print("  %d righe lette." % len(rows))

    contatori = importa(db, rows, clear=args.clear)
    stampa_riepilogo(contatori)
    print("\nImportazione completata.")


if __name__ == '__main__':
    main()
