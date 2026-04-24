#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Importa dati vendite/movimenti da file DELRAI nel database vendite.

Supporta: Messaggerie (xls), Bookwire (csv), KDP Royalties (xlsx),
          Inventario (xlsx), Fiere (xlsx), Ecommerce (xlsx).

Usage:
    python scripts/import_vendite.py --instance cedipg [--clear] [--canale CANALE] [--anno ANNO]

Canali: messaggerie, bookwire, kdp, inventario, fiere, ecommerce, tutti
"""

import sys
import os
import re
import csv
import argparse
from datetime import datetime, date
from decimal import Decimal
from collections import defaultdict
from pathlib import Path

import openpyxl
import xlrd
from gnr.app.gnrapp import GnrApp

PROJECT_ROOT = Path(__file__).parent.parent
DELRAI = PROJECT_ROOT / 'DELRAI'


def parse_decimal_it(val):
    """Converte stringa con virgola decimale in Decimal."""
    if not val:
        return None
    return Decimal(str(val).replace('.', '').replace(',', '.'))


def parse_date_yyyymmdd(val):
    """Converte stringa YYYYMMDD in date."""
    if not val:
        return None
    s = str(val).strip()
    if len(s) == 8:
        return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    return None


def parse_date_ddmmyyyy(val):
    """Converte stringa DD.MM.YYYY in date."""
    if not val:
        return None
    parts = str(val).strip().split('.')
    if len(parts) == 3:
        return date(int(parts[2]), int(parts[1]), int(parts[0]))
    return None


def crea_canali(db):
    """Crea i canali di default se non esistono."""
    tbl = db.table('vendite.canale')
    canali = [
        ('MSG', 'Messaggerie'),
        ('BW', 'Bookwire'),
        ('KDP', 'Amazon KDP'),
        ('BR', 'BookRepublic'),
        ('ECOM', 'Ecommerce'),
        ('FIERA', 'Fiere'),
        ('INV', 'Inventario'),
    ]
    existing = {r['codice'] for r in tbl.query(columns='$codice').fetch()}
    for codice, descrizione in canali:
        if codice not in existing:
            tbl.insert(tbl.newrecord(codice=codice, descrizione=descrizione))
    db.commit()


def build_isbn_titolo_map(db):
    """Costruisce mappa isbn -> titolo_id dal catalogo cedi."""
    tbl_codifica = db.table('cedi.titolo_codifica')
    tbl_titolo = db.table('cedi.titolo')
    isbn_map = {}
    for r in tbl_codifica.query(columns='$codice, $titolo_id').fetch():
        isbn_map[str(r['codice']).strip()] = r['titolo_id']
    for r in tbl_titolo.query(columns='$id, $isbn', where='$isbn IS NOT NULL').fetch():
        isbn_map[str(r['isbn']).strip()] = r['id']
    return isbn_map


def crea_movimento(db, canale_codice, periodo_da, periodo_a, tipo_movimento,
                   file_sorgente, valuta='EUR'):
    """Crea un record testata movimento."""
    tbl = db.table('vendite.movimento')
    rec = tbl.newrecord(
        canale_codice=canale_codice,
        periodo_da=periodo_da,
        periodo_a=periodo_a,
        tipo_movimento=tipo_movimento,
        descrizione='%s %s %s' % (canale_codice, tipo_movimento,
                                   periodo_da.strftime('%Y-%m')),
        file_sorgente=file_sorgente,
        data_importazione=datetime.now(),
        valuta=valuta
    )
    tbl.insert(rec)
    return rec['id']


# ============================================================
# MESSAGGERIE (xls)
# ============================================================
def import_messaggerie_file(db, filepath, isbn_map, tipo):
    """Importa un file Messaggerie vendite o resi."""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    if ws.nrows < 2:
        return 0
    headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
    periodo_da_str = ws.cell_value(1, headers.index('DATA INIZIO PERIODO'))
    periodo_a_str = ws.cell_value(1, headers.index('DATA FINE PERIODO'))
    periodo_da = parse_date_yyyymmdd(periodo_da_str)
    periodo_a = parse_date_yyyymmdd(periodo_a_str)
    if not periodo_da or not periodo_a:
        return 0
    tipo_mov = 'VENDITA' if tipo == 'vendite' else 'RESO'
    mov_id = crea_movimento(db, 'MSG', periodo_da, periodo_a, tipo_mov,
                            os.path.basename(filepath))
    tbl_riga = db.table('vendite.movimento_riga')
    count = 0
    for i in range(1, ws.nrows):
        row = {headers[c]: ws.cell_value(i, c) for c in range(ws.ncols)}
        isbn = str(row.get('CODICE TITOLO ISBN 13 pos', '')).strip()
        if not isbn:
            continue
        quantita = int(row.get('COPIE NETTE', 0) or 0)
        if tipo == 'resi':
            quantita = -abs(quantita)
        rec = tbl_riga.newrecord(
            movimento_id=mov_id,
            titolo_id=isbn_map.get(isbn),
            isbn=isbn,
            descrizione_titolo=str(row.get('DESCRIZIONE CODICE ISBN 13 POSIZIONI', '')).strip(),
            quantita=quantita,
            prezzo_unitario=row.get('PREZZO UNITARIO') or None,
            importo_lordo=row.get('IMPORTO LORDO') or None,
            importo_netto=row.get('NETTO') or None,
            valuta='EUR',
            sconto_percentuale=row.get('SCONTO') or None,
            aliquota_iva=row.get('ALIQUOTA IVA') or None,
            tipo_prodotto='CARTACEO',
            store='Messaggerie'
        )
        tbl_riga.insert(rec)
        count += 1
    totale = sum(r.get('NETTO', 0) or 0
                 for r_idx in range(1, ws.nrows)
                 for r in [{headers[c]: ws.cell_value(r_idx, c) for c in range(ws.ncols)}])
    db.table('vendite.movimento').update(
        dict(id=mov_id, totale_importo=totale, totale_quantita=count),
        dict(id=mov_id)
    )
    return count


def import_messaggerie(db, isbn_map, anno=None):
    """Importa tutti i file Messaggerie."""
    total = 0
    for tipo in ('Vendite', 'Resi'):
        base = DELRAI / '01_Messaggerie' / tipo
        if not base.exists():
            continue
        for year_dir in sorted(base.iterdir()):
            if anno and year_dir.name != str(anno):
                continue
            if not year_dir.is_dir():
                continue
            for f in sorted(year_dir.glob('*.xls')):
                tipo_key = 'vendite' if tipo == 'Vendite' else 'resi'
                n = import_messaggerie_file(db, str(f), isbn_map, tipo_key)
                print("  %s: %d righe" % (f.name, n))
                total += n
                db.commit()
    return total


# ============================================================
# BOOKWIRE (csv)
# ============================================================
def import_bookwire_file(db, filepath, isbn_map):
    """Importa un file Bookwire CSV."""
    with open(filepath, encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        rows = list(reader)
    if not rows:
        return 0
    periodo_da = parse_date_ddmmyyyy(rows[0].get('Period Start Date'))
    periodo_a = parse_date_ddmmyyyy(rows[0].get('Period End Date'))
    if not periodo_da or not periodo_a:
        return 0
    mov_id = crea_movimento(db, 'BW', periodo_da, periodo_a, 'VENDITA',
                            os.path.basename(filepath), valuta='EUR')
    tbl_riga = db.table('vendite.movimento_riga')
    count = 0
    for row in rows:
        isbn = (row.get('ISBN') or '').strip()
        if not isbn:
            continue
        units = int(row.get('Units', 0) or 0)
        rec = tbl_riga.newrecord(
            movimento_id=mov_id,
            titolo_id=isbn_map.get(isbn),
            isbn=isbn,
            descrizione_titolo=(row.get('Title') or '').strip(),
            autore=(row.get('Author') or '').strip(),
            quantita=units,
            prezzo_unitario=parse_decimal_it(row.get('Net Retail Price (Payment Currency)')),
            importo_netto=parse_decimal_it(row.get('Payment Amount Publisher')),
            royalty=parse_decimal_it(row.get('Bookwire Income (Payment Currency)')),
            valuta=row.get('Payment Currency', 'EUR'),
            tasso_cambio=parse_decimal_it(row.get('Exchange Rate')),
            paese=(row.get('Sale Country') or '').strip(),
            store=(row.get('Shop') or '').strip(),
            tipo_prodotto='EBOOK'
        )
        tbl_riga.insert(rec)
        count += 1
    return count


def import_bookwire(db, isbn_map, anno=None):
    """Importa tutti i file Bookwire."""
    total = 0
    base = DELRAI / '02_Bookwire'
    if not base.exists():
        return 0
    for year_dir in sorted(base.iterdir()):
        if anno and year_dir.name != str(anno):
            continue
        if not year_dir.is_dir():
            continue
        for f in sorted(year_dir.glob('*.csv')):
            n = import_bookwire_file(db, str(f), isbn_map)
            print("  %s: %d righe" % (f.name, n))
            total += n
            db.commit()
    return total


# ============================================================
# KDP ROYALTIES (xlsx)
# ============================================================
def import_kdp_file(db, filepath, isbn_map):
    """Importa un file KDP Royalties."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    # trova header (di solito row 2)
    header_row = None
    headers = []
    for i in range(1, min(5, ws.max_row + 1)):
        vals = [c.value for c in ws[i]]
        if any(v and 'Titolo' in str(v) for v in vals):
            header_row = i
            headers = [str(v or '') for v in vals]
            break
    if not header_row:
        return 0
    data_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if any(v for v in row):
            data_rows.append(dict(zip(headers, row)))
    if not data_rows:
        return 0
    # periodo dal nome file: "KDP_Prior_Month_Royalties-2025-01-01-uuid.xlsx"
    # oppure formato vecchio: "KDP Prior Month Royalties-2017-01-timestamp-hash.xlsx"
    fname = os.path.basename(filepath)
    try:
        # cerca pattern YYYY-MM-DD nel nome file
        match = re.search(r'(\d{4})-(\d{2})-(\d{2})', fname)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            periodo_da = date(year, month, 1)
            if month == 12:
                periodo_a = date(year, 12, 31)
            else:
                periodo_a = date(year, month + 1, 1)
        else:
            raise ValueError('No date in filename')
    except (ValueError, OverflowError):
        periodo_da = date.today().replace(day=1)
        periodo_a = periodo_da
    mov_id = crea_movimento(db, 'KDP', periodo_da, periodo_a, 'ROYALTY',
                            fname, valuta='EUR')
    tbl_riga = db.table('vendite.movimento_riga')
    count = 0
    for row in data_rows:
        asin = str(row.get('Codice ASIN', '') or '').strip()
        if not asin:
            continue
        rec = tbl_riga.newrecord(
            movimento_id=mov_id,
            titolo_id=isbn_map.get(asin),
            isbn=asin,
            descrizione_titolo=str(row.get('Titolo', '') or '').strip(),
            autore=str(row.get('Autore', '') or '').strip(),
            quantita=int(row.get('Netto unità vendute', 0) or 0),
            royalty=row.get('Royalty') or None,
            valuta=str(row.get('Valuta', 'EUR') or 'EUR').strip(),
            prezzo_unitario=row.get('Prezzo medio di listino escluse le imposte'),
            paese=str(row.get('Mercato', '') or '').strip(),
            store='Amazon KDP',
            tipo_prodotto='EBOOK'
        )
        tbl_riga.insert(rec)
        count += 1
    return count


def import_kdp(db, isbn_map, anno=None):
    """Importa tutti i file KDP Royalties."""
    total = 0
    base = DELRAI / '03_KDP' / 'Royalties'
    if not base.exists():
        return 0
    for year_dir in sorted(base.iterdir()):
        if anno and year_dir.name != str(anno):
            continue
        if not year_dir.is_dir():
            continue
        for f in sorted(year_dir.glob('*.xlsx')):
            n = import_kdp_file(db, str(f), isbn_map)
            print("  %s: %d righe" % (f.name, n))
            total += n
            db.commit()
    return total


# ============================================================
# INVENTARIO / FIERE / ECOMMERCE (xlsx semplice)
# ============================================================
def import_xlsx_semplice(db, filepath, isbn_map, tipo, canale_codice):
    """Importa file xlsx semplice (inventario/fiere/ecommerce)."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [str(c.value or '') for c in ws[1]]
    tbl = db.table('vendite.inventario')
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        record = dict(zip(headers, row))
        isbn = str(record.get('ISBN', '') or '').strip()
        if not isbn:
            continue
        data_val = record.get('Data')
        if isinstance(data_val, datetime):
            data_val = data_val.date()
        quantita = record.get('Quantità') or record.get('Omaggio') or 0
        rec = tbl.newrecord(
            titolo_id=isbn_map.get(isbn),
            isbn=isbn,
            data=data_val,
            quantita=int(quantita) if quantita else 0,
            ricavo=record.get('Ricavo') or None,
            tipo=tipo.upper(),
            note=str(record.get('Note', '') or '').strip() or os.path.basename(filepath)
        )
        tbl.insert(rec)
        count += 1
    return count


def import_inventario(db, isbn_map, anno=None):
    """Importa tutti i file inventario."""
    total = 0
    base = DELRAI / '04_Storico' / 'Inventario'
    if not base.exists():
        return 0
    for year_dir in sorted(base.iterdir()):
        if anno and year_dir.name != str(anno):
            continue
        if not year_dir.is_dir():
            continue
        for f in sorted(year_dir.glob('*.xlsx')):
            n = import_xlsx_semplice(db, str(f), isbn_map, 'CONCORSO', 'INV')
            print("  %s: %d righe" % (f.name, n))
            total += n
            db.commit()
    return total


def import_fiere(db, isbn_map, anno=None):
    """Importa tutti i file fiere."""
    total = 0
    base = DELRAI / '06_Fiere'
    if not base.exists():
        return 0
    for year_dir in sorted(base.iterdir()):
        if anno and year_dir.name != str(anno):
            continue
        if not year_dir.is_dir():
            continue
        for f in sorted(year_dir.glob('*.xlsx')):
            n = import_xlsx_semplice(db, str(f), isbn_map, 'FIERA', 'FIERA')
            print("  %s: %d righe" % (f.name, n))
            total += n
            db.commit()
    return total


def import_ecommerce(db, isbn_map, anno=None):
    """Importa tutti i file ecommerce."""
    total = 0
    base = DELRAI / '07_Ecommerce'
    if not base.exists():
        return 0
    for year_dir in sorted(base.iterdir()):
        if anno and year_dir.name != str(anno):
            continue
        if not year_dir.is_dir():
            continue
        for f in sorted(year_dir.glob('*.xlsx')):
            n = import_xlsx_semplice(db, str(f), isbn_map, 'ECOMMERCE', 'ECOM')
            print("  %s: %d righe" % (f.name, n))
            total += n
            db.commit()
    return total


# ============================================================
# MAIN
# ============================================================
def clear_vendite(db):
    """Svuota tutte le tabelle vendite."""
    print("Pulizia tabelle vendite...")
    db.rollback()
    for tbl_name in ('vendite_movimento_riga', 'vendite_movimento', 'vendite_inventario'):
        db.execute('TRUNCATE TABLE vendite.%s CASCADE' % tbl_name)
    db.commit()
    print("  Tabelle svuotate.")


def stampa_riepilogo(contatori):
    print("\n" + "=" * 50)
    print("RIEPILOGO IMPORTAZIONE VENDITE")
    print("=" * 50)
    for canale, count in contatori.items():
        print("  %-20s: %d righe" % (canale, count))
    print("  %-20s: %d righe" % ('TOTALE', sum(contatori.values())))
    print("=" * 50)


def main():
    parser = argparse.ArgumentParser(description='Importa vendite/movimenti da DELRAI')
    parser.add_argument('--instance', default='cedipg', help='Nome istanza')
    parser.add_argument('--clear', action='store_true', help='Svuota tabelle prima di importare')
    parser.add_argument('--canale', default='tutti',
                        help='Canale: messaggerie, bookwire, kdp, inventario, fiere, ecommerce, tutti')
    parser.add_argument('--anno', type=int, default=None, help='Filtra per anno')
    args = parser.parse_args()

    print("Connessione a istanza: %s" % args.instance)
    instance_dir = PROJECT_ROOT / 'instances' / args.instance
    if not instance_dir.exists():
        print("Errore: directory istanza non trovata: %s" % instance_dir)
        sys.exit(1)

    app = GnrApp(str(instance_dir))
    db = app.db

    print("Verifica schema database...")
    db.checkDb(applyChanges=True)

    crea_canali(db)

    if args.clear:
        clear_vendite(db)

    isbn_map = build_isbn_titolo_map(db)
    print("Mappa ISBN/ASIN: %d codici caricati dal catalogo." % len(isbn_map))

    contatori = {}
    canale = args.canale.lower()

    if canale in ('messaggerie', 'tutti'):
        print("\n--- Messaggerie ---")
        contatori['Messaggerie'] = import_messaggerie(db, isbn_map, args.anno)

    if canale in ('bookwire', 'tutti'):
        print("\n--- Bookwire ---")
        contatori['Bookwire'] = import_bookwire(db, isbn_map, args.anno)

    if canale in ('kdp', 'tutti'):
        print("\n--- KDP Royalties ---")
        contatori['KDP'] = import_kdp(db, isbn_map, args.anno)

    if canale in ('inventario', 'tutti'):
        print("\n--- Inventario ---")
        contatori['Inventario'] = import_inventario(db, isbn_map, args.anno)

    if canale in ('fiere', 'tutti'):
        print("\n--- Fiere ---")
        contatori['Fiere'] = import_fiere(db, isbn_map, args.anno)

    if canale in ('ecommerce', 'tutti'):
        print("\n--- Ecommerce ---")
        contatori['Ecommerce'] = import_ecommerce(db, isbn_map, args.anno)

    stampa_riepilogo(contatori)
    print("\nImportazione completata.")


if __name__ == '__main__':
    main()
